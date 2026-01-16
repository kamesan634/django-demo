"""
Export utilities for generating CSV, Excel, and PDF reports.
"""
import csv
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse


class ExportService:
    """Service for exporting data to various formats."""

    @staticmethod
    def to_csv(data, filename, columns=None):
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries
            filename: Output filename (without extension)
            columns: List of column definitions [(key, header), ...]
                     If None, uses dict keys as both key and header
        """
        response = HttpResponse(
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        if not data:
            return response

        # Determine columns
        if columns is None:
            columns = [(k, k) for k in data[0].keys()]

        writer = csv.writer(response)

        # Write header
        writer.writerow([col[1] for col in columns])

        # Write data
        for row in data:
            writer.writerow([
                ExportService._format_value(row.get(col[0], ''))
                for col in columns
            ])

        return response

    @staticmethod
    def to_excel(data, filename, columns=None, sheet_name='Sheet1'):
        """
        Export data to Excel format.

        Args:
            data: List of dictionaries
            filename: Output filename (without extension)
            columns: List of column definitions [(key, header), ...]
            sheet_name: Excel sheet name
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return ExportService.to_csv(data, filename, columns)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        if not data:
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response

        # Determine columns
        if columns is None:
            columns = [(k, k) for k in data[0].keys()]

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color='DDEEFF',
            end_color='DDEEFF',
            fill_type='solid'
        )

        # Write header
        for col_idx, (_, header) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write data
        for row_idx, row in enumerate(data, start=2):
            for col_idx, (key, _) in enumerate(columns, start=1):
                value = ExportService._format_value(row.get(key, ''))
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, (key, header) in enumerate(columns, start=1):
            max_length = len(str(header))
            for row in data:
                value = str(row.get(key, ''))
                max_length = max(max_length, len(value))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    @staticmethod
    def to_pdf(data, filename, columns=None, title=None):
        """
        Export data to PDF format.

        Args:
            data: List of dictionaries
            filename: Output filename (without extension)
            columns: List of column definitions [(key, header), ...]
            title: Report title
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            # Fallback to CSV if reportlab not installed
            return ExportService.to_csv(data, filename, columns)

        buffer = io.BytesIO()

        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        elements = []
        styles = getSampleStyleSheet()

        # Title
        if title:
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=12
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 12))

        # Determine columns
        if columns is None and data:
            columns = [(k, k) for k in data[0].keys()]

        if data and columns:
            # Table data
            table_data = [[col[1] for col in columns]]

            for row in data:
                table_data.append([
                    str(ExportService._format_value(row.get(col[0], '')))
                    for col in columns
                ])

            # Calculate column widths
            page_width = landscape(A4)[0] - 60
            col_width = page_width / len(columns)

            table = Table(table_data, colWidths=[col_width] * len(columns))

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DDEEFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
            ]))

            elements.append(table)

        # Timestamp
        elements.append(Spacer(1, 12))
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Paragraph(f'Generated: {timestamp}', styles['Normal']))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    @staticmethod
    def _format_value(value):
        """Format value for export."""
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        return value
