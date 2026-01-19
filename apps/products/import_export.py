"""
Product import/export services.
F03-005: 商品匯入匯出
"""
import io
import csv
import logging
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Any, Tuple, Optional

from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)


class ProductImportService:
    """
    Service for importing products from Excel/CSV files.
    F03-005: 商品匯入匯出
    """

    # Required and optional columns
    REQUIRED_COLUMNS = ['sku', 'name', 'sale_price']
    OPTIONAL_COLUMNS = [
        'category', 'description', 'cost_price', 'unit',
        'tax_type', 'safety_stock', 'status', 'barcode'
    ]

    # Column mapping (file header -> model field)
    COLUMN_MAPPING = {
        '商品編號': 'sku',
        'SKU': 'sku',
        'sku': 'sku',
        '商品名稱': 'name',
        '名稱': 'name',
        'name': 'name',
        '分類': 'category',
        'category': 'category',
        '描述': 'description',
        'description': 'description',
        '售價': 'sale_price',
        '銷售價格': 'sale_price',
        'sale_price': 'sale_price',
        '成本': 'cost_price',
        '成本價': 'cost_price',
        'cost_price': 'cost_price',
        '單位': 'unit',
        'unit': 'unit',
        '稅別': 'tax_type',
        'tax_type': 'tax_type',
        '安全庫存': 'safety_stock',
        'safety_stock': 'safety_stock',
        '狀態': 'status',
        'status': 'status',
        '條碼': 'barcode',
        'barcode': 'barcode',
    }

    @classmethod
    def import_from_excel(cls, file, user=None, update_existing=False) -> Dict:
        """
        Import products from Excel file.

        Args:
            file: Uploaded Excel file
            user: User performing the import
            update_existing: Whether to update existing products

        Returns:
            Dict with import results
        """
        try:
            import openpyxl
        except ImportError:
            return {
                'success': False,
                'message': '缺少 openpyxl 套件，請安裝後再試',
                'created': 0,
                'updated': 0,
                'errors': []
            }

        try:
            workbook = openpyxl.load_workbook(file, read_only=True)
            sheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            # Map headers to field names
            field_mapping = cls._map_headers(headers)

            # Validate required columns
            missing = cls._validate_required_columns(field_mapping)
            if missing:
                return {
                    'success': False,
                    'message': f'缺少必要欄位: {", ".join(missing)}',
                    'created': 0,
                    'updated': 0,
                    'errors': []
                }

            # Process rows
            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {}
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx] in field_mapping:
                        field_name = field_mapping[headers[col_idx]]
                        row_data[field_name] = cell_value
                row_data['_row'] = row_idx
                rows.append(row_data)

            workbook.close()

            # Import products
            return cls._import_products(rows, user, update_existing)

        except Exception as e:
            logger.error(f"Excel import error: {e}")
            return {
                'success': False,
                'message': f'匯入失敗: {str(e)}',
                'created': 0,
                'updated': 0,
                'errors': []
            }

    @classmethod
    def import_from_csv(cls, file, user=None, update_existing=False) -> Dict:
        """
        Import products from CSV file.

        Args:
            file: Uploaded CSV file
            user: User performing the import
            update_existing: Whether to update existing products

        Returns:
            Dict with import results
        """
        try:
            # Read CSV content
            content = file.read()
            if isinstance(content, bytes):
                # Try UTF-8 first, then Big5 for traditional Chinese
                try:
                    content = content.decode('utf-8-sig')
                except UnicodeDecodeError:
                    content = content.decode('big5')

            reader = csv.DictReader(io.StringIO(content))

            # Map headers
            field_mapping = cls._map_headers(reader.fieldnames)

            # Validate required columns
            missing = cls._validate_required_columns(field_mapping)
            if missing:
                return {
                    'success': False,
                    'message': f'缺少必要欄位: {", ".join(missing)}',
                    'created': 0,
                    'updated': 0,
                    'errors': []
                }

            # Process rows
            rows = []
            for row_idx, row in enumerate(reader, start=2):
                row_data = {}
                for header, value in row.items():
                    if header in field_mapping:
                        row_data[field_mapping[header]] = value
                row_data['_row'] = row_idx
                rows.append(row_data)

            # Import products
            return cls._import_products(rows, user, update_existing)

        except Exception as e:
            logger.error(f"CSV import error: {e}")
            return {
                'success': False,
                'message': f'匯入失敗: {str(e)}',
                'created': 0,
                'updated': 0,
                'errors': []
            }

    @classmethod
    def _map_headers(cls, headers) -> Dict[str, str]:
        """Map file headers to model field names."""
        mapping = {}
        for header in headers:
            if header and header in cls.COLUMN_MAPPING:
                mapping[header] = cls.COLUMN_MAPPING[header]
        return mapping

    @classmethod
    def _validate_required_columns(cls, field_mapping: Dict) -> List[str]:
        """Validate that all required columns are present."""
        mapped_fields = set(field_mapping.values())
        missing = []
        for req in cls.REQUIRED_COLUMNS:
            if req not in mapped_fields:
                missing.append(req)
        return missing

    @classmethod
    @transaction.atomic
    def _import_products(cls, rows: List[Dict], user, update_existing: bool) -> Dict:
        """Import products from parsed rows."""
        from apps.products.models import Product, Category, Unit, TaxType, ProductBarcode

        created_count = 0
        updated_count = 0
        errors = []

        # Cache lookups
        categories = {c.name: c for c in Category.objects.filter(is_deleted=False)}
        units = {u.name: u for u in Unit.objects.filter(is_deleted=False)}
        tax_types = {t.name: t for t in TaxType.objects.filter(is_deleted=False)}

        for row in rows:
            row_num = row.pop('_row', '?')

            try:
                # Validate and clean data
                sku = str(row.get('sku', '')).strip()
                name = str(row.get('name', '')).strip()

                if not sku:
                    errors.append({'row': row_num, 'error': 'SKU 不能為空'})
                    continue
                if not name:
                    errors.append({'row': row_num, 'error': '商品名稱不能為空'})
                    continue

                # Parse sale_price
                try:
                    sale_price = Decimal(str(row.get('sale_price', 0)))
                except (InvalidOperation, ValueError):
                    errors.append({'row': row_num, 'error': '售價格式錯誤'})
                    continue

                # Check if product exists
                existing = Product.objects.filter(sku=sku).first()

                if existing:
                    if not update_existing:
                        errors.append({'row': row_num, 'error': f'SKU {sku} 已存在'})
                        continue

                    # Update existing product
                    existing.name = name
                    existing.sale_price = sale_price
                    cls._update_product_fields(existing, row, categories, units, tax_types)
                    existing.updated_by = user
                    existing.save()
                    updated_count += 1
                else:
                    # Create new product
                    product = Product(
                        sku=sku,
                        name=name,
                        sale_price=sale_price,
                        created_by=user
                    )
                    cls._update_product_fields(product, row, categories, units, tax_types)
                    product.save()

                    # Create barcode if provided
                    barcode = str(row.get('barcode', '')).strip()
                    if barcode:
                        ProductBarcode.objects.create(
                            product=product,
                            barcode=barcode,
                            barcode_type='CUSTOM',
                            is_primary=True,
                            created_by=user
                        )

                    created_count += 1

            except Exception as e:
                logger.error(f"Error importing row {row_num}: {e}")
                errors.append({'row': row_num, 'error': str(e)})

        return {
            'success': True,
            'message': f'匯入完成：新增 {created_count} 筆，更新 {updated_count} 筆',
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        }

    @classmethod
    def _update_product_fields(cls, product, row: Dict, categories: Dict, units: Dict, tax_types: Dict):
        """Update product fields from row data."""
        # Category
        category_name = str(row.get('category', '')).strip()
        if category_name and category_name in categories:
            product.category = categories[category_name]

        # Description
        if 'description' in row:
            product.description = str(row.get('description', '')).strip()

        # Cost price
        if row.get('cost_price'):
            try:
                product.cost_price = Decimal(str(row['cost_price']))
            except (InvalidOperation, ValueError):
                pass

        # Unit
        unit_name = str(row.get('unit', '')).strip()
        if unit_name and unit_name in units:
            product.unit = units[unit_name]

        # Tax type
        tax_name = str(row.get('tax_type', '')).strip()
        if tax_name and tax_name in tax_types:
            product.tax_type = tax_types[tax_name]

        # Safety stock
        if row.get('safety_stock'):
            try:
                product.safety_stock = int(row['safety_stock'])
            except ValueError:
                pass

        # Status
        status = str(row.get('status', '')).strip().upper()
        if status in ['ACTIVE', 'INACTIVE', 'DISCONTINUED']:
            product.status = status


class ProductExportService:
    """
    Service for exporting products to Excel/CSV files.
    F03-005: 商品匯入匯出
    """

    # Default export columns
    DEFAULT_COLUMNS = [
        ('sku', '商品編號'),
        ('name', '商品名稱'),
        ('category__name', '分類'),
        ('sale_price', '售價'),
        ('cost_price', '成本價'),
        ('unit__name', '單位'),
        ('tax_type__name', '稅別'),
        ('safety_stock', '安全庫存'),
        ('status', '狀態'),
        ('description', '描述'),
    ]

    @classmethod
    def export_to_excel(cls, queryset=None, columns=None):
        """
        Export products to Excel file.

        Args:
            queryset: Product queryset (optional, defaults to all active)
            columns: List of (field, header) tuples

        Returns:
            HttpResponse with Excel file
        """
        from apps.core.export import ExportService
        from apps.products.models import Product

        if queryset is None:
            queryset = Product.objects.filter(
                is_deleted=False
            ).select_related('category', 'unit', 'tax_type')

        if columns is None:
            columns = cls.DEFAULT_COLUMNS

        # Get data
        data = cls._prepare_export_data(queryset)

        filename = f'products_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        return ExportService.to_excel(data, filename, columns, '商品清單')

    @classmethod
    def export_to_csv(cls, queryset=None, columns=None):
        """
        Export products to CSV file.

        Args:
            queryset: Product queryset
            columns: List of (field, header) tuples

        Returns:
            HttpResponse with CSV file
        """
        from apps.core.export import ExportService
        from apps.products.models import Product

        if queryset is None:
            queryset = Product.objects.filter(
                is_deleted=False
            ).select_related('category', 'unit', 'tax_type')

        if columns is None:
            columns = cls.DEFAULT_COLUMNS

        # Get data
        data = cls._prepare_export_data(queryset)

        filename = f'products_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        return ExportService.to_csv(data, filename, columns)

    @classmethod
    def get_template(cls):
        """
        Generate an import template Excel file.

        Returns:
            HttpResponse with template Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from django.http import HttpResponse
        except ImportError:
            from django.http import HttpResponse
            return HttpResponse('缺少 openpyxl 套件', status=500)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = '商品匯入範本'

        # Define headers
        headers = [
            ('商品編號', True),
            ('商品名稱', True),
            ('分類', False),
            ('售價', True),
            ('成本價', False),
            ('單位', False),
            ('稅別', False),
            ('安全庫存', False),
            ('狀態', False),
            ('條碼', False),
            ('描述', False),
        ]

        # Style definitions
        header_font = Font(bold=True, color='FFFFFF')
        required_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        optional_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, (header, required) in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = required_fill if required else optional_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Add sample data
        sample_data = [
            ['PROD001', '白色T-Shirt', '上衣', 299, 150, '件', '應稅', 50, 'ACTIVE', '4710001234567', '純棉材質'],
            ['PROD002', '黑色長褲', '褲子', 599, 300, '件', '應稅', 30, 'ACTIVE', '4710001234568', '彈性布料'],
        ]

        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths
        column_widths = [15, 20, 10, 10, 10, 8, 10, 10, 10, 15, 30]
        for col, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Add instructions sheet
        instructions = workbook.create_sheet(title='說明')
        instructions_data = [
            ['欄位說明'],
            [''],
            ['必填欄位 (藍色):'],
            ['- 商品編號: 唯一識別碼，不可重複'],
            ['- 商品名稱: 商品的顯示名稱'],
            ['- 售價: 銷售價格，數字格式'],
            [''],
            ['選填欄位 (綠色):'],
            ['- 分類: 需與系統中的分類名稱完全一致'],
            ['- 成本價: 成本價格，數字格式'],
            ['- 單位: 需與系統中的單位名稱完全一致'],
            ['- 稅別: 需與系統中的稅別名稱完全一致'],
            ['- 安全庫存: 整數'],
            ['- 狀態: ACTIVE (銷售中), INACTIVE (停售), DISCONTINUED (已下架)'],
            ['- 條碼: 商品條碼'],
            ['- 描述: 商品描述文字'],
        ]

        for row_idx, row_data in enumerate(instructions_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                instructions.cell(row=row_idx, column=col_idx, value=value)

        instructions.column_dimensions['A'].width = 60

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
        workbook.save(response)

        return response

    @classmethod
    def _prepare_export_data(cls, queryset) -> List[Dict]:
        """Prepare data for export."""
        data = []
        for product in queryset:
            data.append({
                'sku': product.sku,
                'name': product.name,
                'category__name': product.category.name if product.category else '',
                'sale_price': float(product.sale_price),
                'cost_price': float(product.cost_price),
                'unit__name': product.unit.name if product.unit else '',
                'tax_type__name': product.tax_type.name if product.tax_type else '',
                'safety_stock': product.safety_stock,
                'status': product.status,
                'description': product.description or '',
            })
        return data
